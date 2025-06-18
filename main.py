import tkinter
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os
import uuid
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import datetime

FILENAME = "adatok.xlsx"
APPOINTMENT_FILE = "idopontok.xlsx"

# --- Excel kezelő függvények ---
def save_to_excel(patient, filename=FILENAME):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Név", "Szul. dátum", "Telefon", "Email"])
    ws.append([patient["ID"], patient["Név"], patient["Szul. dátum"], patient["Telefon"], patient["Email"]])
    wb.save(filename)

def load_patients(filename=FILENAME):
    patients = []
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                patients.append({
                    "ID": row[0],
                    "Név": row[1],
                    "Szul. dátum": row[2],
                    "Telefon": row[3],
                    "Email": row[4]
                })
    return patients

def save_all_patients(patients, filename=FILENAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Név", "Szul. dátum", "Telefon", "Email"])
    for p in patients:
        ws.append([p["ID"], p["Név"], p["Szul. dátum"], p["Telefon"], p["Email"]])
    wb.save(filename)

from tkinter.filedialog import asksaveasfilename

def export_patient_to_pdf(patient):
    default_filename = f"{patient['Név'].replace(' ', '_')}_export.pdf"
    filename = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF fájlok", "*.pdf")], initialfile=default_filename)
    if not filename:
        return
    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4
    y = height - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, f"Páciens adatlap: {patient['Név']}")
    y -= 30
    c.setFont("Helvetica", 10)
    for line in [
        f"Név: {patient['Név']}",
        f"Email: {patient['Email']}",
        f"Születési dátum: {patient['Szul. dátum']}",
        f"Telefon: {patient['Telefon']}",
        f"ID: {patient['ID']}"
    ]:
        c.drawString(50, y, line)
        y -= 15
    c.save()

# --- Időpont foglalás mentés ---
def save_appointment(patient_id, patient_name, date, time, note):
    if os.path.exists(APPOINTMENT_FILE):
        wb = load_workbook(APPOINTMENT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Páciens ID", "Név", "Dátum", "Időpont", "Megjegyzés"])
    ws.append([patient_id, patient_name, date, time, note])
    wb.save(APPOINTMENT_FILE)

# --- Alkalmazás GUI ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Masszőr Program")
        self.geometry("1000x700")

        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.refresh_patients_list())
        search_entry = ctk.CTkEntry(self, textvariable=self.search_var, placeholder_text="Keresés név vagy email szerint...")
        search_entry.pack(pady=10, padx=20, fill="x")

        self.patients_frame = ctk.CTkFrame(self)
        self.patients_frame.pack(pady=10, fill="both", expand=True)

        control_frame = ctk.CTkFrame(self)
        control_frame.pack(pady=10)
        ctk.CTkButton(control_frame, text="Új páciens", command=self.open_edit_popup).pack(side="left", padx=10)

        self.refresh_patients_list()

    def refresh_patients_list(self):
        for widget in self.patients_frame.winfo_children():
            widget.destroy()

        query = self.search_var.get().lower()
        self.patients = [p for p in load_patients() if query in p["Név"].lower() or query in p["Email"].lower()]

        for patient in self.patients:
            row = ctk.CTkFrame(self.patients_frame)
            row.pack(fill="x", padx=10, pady=3)

            label = ctk.CTkLabel(row, text=f"{patient['Név']}  <{patient['Email']}>", anchor="w")
            label.pack(side="left", padx=10, fill="x", expand=True)

            ctk.CTkButton(row, text="PDF", width=60, command=lambda p=patient: export_patient_to_pdf(p)).pack(side="right", padx=5)
            ctk.CTkButton(row, text="Megnyitás", width=100, command=lambda p=patient: self.open_patient_detail(p)).pack(side="right", padx=5)

    def open_patient_detail(self, patient):
        popup = ctk.CTkToplevel(self)
        popup.title(f"{patient['Név']} - Adatlap")
        popup.geometry("400x350")

        for key, val in patient.items():
            ctk.CTkLabel(popup, text=f"{key}: {val}").pack(pady=5)

        ctk.CTkButton(popup, text="Szerkesztés", command=lambda: [popup.destroy(), self.open_edit_popup(patient)]).pack(pady=5)
        ctk.CTkButton(popup, text="Törlés", fg_color="red", command=lambda: [popup.destroy(), self.delete_patient(patient)]).pack(pady=5)
        ctk.CTkButton(popup, text="Időpont foglalás", command=lambda: [popup.destroy(), self.book_appointment(patient)]).pack(pady=5)

    def delete_patient(self, patient):
        self.patients = [p for p in load_patients() if p["ID"] != patient["ID"]]
        save_all_patients(self.patients)
        self.refresh_patients_list()

    def open_edit_popup(self, patient=None):
        edit = ctk.CTkToplevel(self)
        edit.geometry("400x400")
        edit.title("Páciens szerkesztése" if patient else "Új páciens")

        name_var = ctk.StringVar(value=patient["Név"] if patient else "")
        dob_var = ctk.StringVar(value=patient["Szul. dátum"] if patient else "")
        phone_var = ctk.StringVar(value=patient["Telefon"] if patient else "")
        email_var = ctk.StringVar(value=patient["Email"] if patient else "")

        for label, var in [
            ("Név", name_var),
            ("Szul. dátum", dob_var),
            ("Telefon", phone_var),
            ("Email", email_var)
        ]:
            ctk.CTkLabel(edit, text=label).pack()
            ctk.CTkEntry(edit, textvariable=var).pack(pady=5)

        def save():
            new_data = {
                "ID": patient["ID"] if patient else str(uuid.uuid4()),
                "Név": name_var.get(),
                "Szul. dátum": dob_var.get(),
                "Telefon": phone_var.get(),
                "Email": email_var.get()
            }
            data = [new_data if p["ID"] == new_data["ID"] else p for p in load_patients()]
            if not patient:
                data.append(new_data)
            save_all_patients(data)
            edit.destroy()
            self.refresh_patients_list()

        ctk.CTkButton(edit, text="Mentés", command=save).pack(pady=10)

    def book_appointment(self, patient):
        book = ctk.CTkToplevel(self)
        book.geometry("400x300")
        book.title("Időpont foglalás")

        date_var = ctk.StringVar()
        time_var = ctk.StringVar()
        note_var = ctk.StringVar()

        for label, var in [
            ("Dátum (YYYY-MM-DD)", date_var),
            ("Időpont (pl. 14:00)", time_var),
            ("Megjegyzés", note_var)
        ]:
            ctk.CTkLabel(book, text=label).pack()
            ctk.CTkEntry(book, textvariable=var).pack(pady=5)

        def save_appt():
            save_appointment(patient["ID"], patient["Név"], date_var.get(), time_var.get(), note_var.get())
            book.destroy()

        ctk.CTkButton(book, text="Foglalás mentése", command=save_appt).pack(pady=10)

if __name__ == "__main__":
    ctk.set_default_color_theme("dark-blue")
    ctk.set_appearance_mode("system")
    app = App()
    app.mainloop()
