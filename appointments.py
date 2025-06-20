import tkinter
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os
import uuid
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import datetime, date
from tkcalendar import DateEntry
from tkinter import ttk
from tkinter.filedialog import asksaveasfilename
from tkinter import ttk

FILENAME = "data/adatok.xlsx"
APPOINTMENT_FILE = "data/idopontok.xlsx"


def save_appointment(patient_id, patient_name, date, time, note):
    
    if os.path.exists(APPOINTMENT_FILE):
        wb = load_workbook(APPOINTMENT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Páciens ID", "Név", "Dátum", "Időpont", "Megjegyzés"])
    ws.append([patient_id, patient_name, date, time, note])
    wb.save(APPOINTMENT_FILE)

def load_appointments():
    appointments = []
    if os.path.exists(APPOINTMENT_FILE):
        wb = load_workbook(APPOINTMENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                appointments.append({
                    "Páciens ID": row[0],
                    "Név": row[1],
                    "Dátum": row[2],
                    "Időpont": row[3],
                    "Megjegyzés": row[4]
                })
    return appointments

def save_all_appointments(appointments):
    wb = Workbook()
    ws = wb.active
    # Fejléc
    headers = ["ID", "Név", "Telefon", "Email", "Szul. dátum", "Dátum", "Időpont", "Megjegyzés"]
    ws.append(headers)
    # Adatok
    for a in appointments:
        ws.append([
            a.get("ID", ""),
            a.get("Név", ""),
            a.get("Telefon", ""),
            a.get("Email", ""),
            a.get("Szul. dátum", ""),
            a.get("Dátum", ""),
            a.get("Időpont", ""),
            a.get("Megjegyzés", "")
        ])
    # Fájl mentése
    os.makedirs(os.path.dirname(APPOINTMENT_FILE), exist_ok=True)
    wb.save(APPOINTMENT_FILE)