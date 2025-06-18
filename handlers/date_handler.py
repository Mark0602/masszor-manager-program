import tkinter
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os
import uuid
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import datetime, date
from tkcalendar import DateEntry
from tkinter import ttk
from tkinter.filedialog import asksaveasfilename
from tkinter import ttk


# --- Excel kezelő függvények ---
def save_to_excel(patient, filename=FILENAME):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Név", "Szul. dátum", "Telefon", "Email"])
    ws.append([patient["ID"], patient["Név"], patient["Szul. dátum"], patient["Telefon"], patient["Email"]])
    wb.save(filename)

def load_patients(filename=FILENAME):
    patients = []
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                patients.append({
                    "ID": row[0],
                    "Név": row[1],
                    "Szul. dátum": row[2],
                    "Telefon": row[3],
                    "Email": row[4]
                })
    return patients

def save_all_patients(patients, filename=FILENAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Név", "Szul. dátum", "Telefon", "Email"])
    for p in patients:
        ws.append([p["ID"], p["Név"], p["Szul. dátum"], p["Telefon"], p["Email"]])
    wb.save(filename)