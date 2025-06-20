import tkinter
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import os
import uuid
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from datetime import datetime, date
from tkcalendar import DateEntry
from tkinter import ttk
from tkinter.filedialog import asksaveasfilename
from tkinter import ttk
FILENAME = "data/adatok.xlsx"
APPOINTMENT_FILE = "data/idopontok.xlsx"


# --- Excel kezelő függvények ---
def save_to_excel(patient, filename=FILENAME):
    headers = ["ID", "Név", "Szul. dátum", "Telefon", "Email"]
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    # Mindig a helyes sorrendben, hiányzó mezőket üres stringgel pótolva
    row = [patient.get(h, "") for h in headers]
    ws.append(row)
    wb.save(filename)

def load_patients(filename=FILENAME):
    patients = []
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                patients.append({
                    "ID": row[0],
                    "Név": row[1],
                    "Szul. dátum": row[2],
                    "Telefon": row[3],
                    "Email": row[4]
                })
    return patients

def save_all_patients(patients, filename=FILENAME):
    headers = ["ID", "Név", "Szul. dátum", "Telefon", "Email"]
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for p in patients:
        row = [p.get(h, "") for h in headers]
        ws.append(row)
    wb.save(filename)